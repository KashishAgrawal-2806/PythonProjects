import openpyxl as xl
from openpyxl.chart import BarChart,Reference

# load the Transtion sheet
wb=xl.load_workbook("Transaction.xlsx")
sheet=wb['Sheet1']
# WAy 1
# cell=sheet['a1']
# Way 2
# cell=sheet.cell(1,1);
# print(sheet.max_row);
for i in range (2,sheet.max_row+1):
    cell=sheet.cell(i,3);
    corect_price=cell.value*0.9;
    corrected_cell=sheet.cell(i,4)
    corrected_cell.value=corect_price;

values=Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save("Trans2.xlsx")